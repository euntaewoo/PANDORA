import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_qoo10_excel():
    # 1. 워크북 및 시트 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "행사별 정산 비교 계산기"
    
    # 격자선 보이게 설정
    ws.views.sheetView[0].showGridLines = True

    # 2. 스타일 정의 (비즈니스 깔끔한 네이비 테마)
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=11)
    font_bold = Font(name="맑은 고딕", size=11, bold=True)
    font_total = Font(name="맑은 고딕", size=11, bold=True, color="9C0006") # 강조용 붉은색계열

    fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # 진한 네이비
    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # 모던 그레이블루
    fill_input = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 입력칸 옅은 노란색
    fill_total = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # 합계칸 옅은 빨간색

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="000000"))
    double_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    # 3. 타이틀 생성
    ws.merge_cells("A1:D1")
    ws["A1"] = "큐텐재팬(Qoo10 Japan) 메가와리 vs 메가포 실무 정산 비교 시뮬레이터"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40

    # 4. 기본 변수 입력 영역 (환율 및 기본수수료율)
    ws["A3"] = "현재 적용 엔화 환율 (1엔당 원화)"
    ws["A3"].font = font_bold
    ws["B3"] = 9.6
    ws["B3"].font = font_bold
    ws["B3"].fill = fill_input
    ws["B3"].number_format = '#,##0.0'
    ws["B3"].alignment = align_right

    ws["A4"] = "기본 카테고리 수수료율"
    ws["A4"].font = font_bold
    ws["B4"] = 0.10
    ws["B4"].font = font_bold
    ws["B4"].fill = fill_input
    ws["B4"].number_format = '0%'
    ws["B4"].alignment = align_right

    ws["A5"] = "메가포 판매자 포인트 분담률"
    ws["A5"].font = font_bold
    ws["B5"] = 0.05
    ws["B5"].font = font_bold
    ws["B5"].fill = fill_input
    ws["B5"].number_format = '0%'
    ws["B5"].alignment = align_right
    ws["C5"] = "(기본 5%, 선택에 따라 10%, 15%, 20% 직접 입력 가능)"
    ws["C5"].font = Font(name="맑은 고딕", size=9, italic=True, color="595959")

    # 5. 테이블 헤더 작성
    headers = ["구분 항목", "1. 메가와리 (MEGA할인)", "2. 메가포 (MEGA POINT)", "비고 및 계산 규칙"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[7].height = 28

    # 6. 데이터 세부 행 내용 및 수식 정의
    # (행 이름, 메가와리 수식/값, 메가포 수식/값, 비고 설명)
    rows_data = [
        ("할인 적용 전 판매가격 (입력)", 5200, 5200, "JQSM에 등록된 상품의 기본 판매가 (노란색 셀에 변경 가능)"),
        ("고객 최종 실결제금액", "=B8*(1-0.2)", "=C8*(1-0.1)", "메가와리 20% 즉시할인 / 메가포 10% 즉시할인 적용액"),
        ("① 기본 카테고리 수수료", "=B8*$B$4", "=C8*$B$4", "할인 적용 전 판매가격 기준 수수료 부과"),
        ("② 판매자 쿠폰 분담금", "=B8*0.10", "=C8*0.05", "메가와리: 쿠폰 20%의 절반(10%) / 메가포: 쿠폰 10%의 절반(5%)"),
        ("③ 판매자 포인트 분담금", 0, "=IF((C9*(0.05+$B$5))>10000, INT(10000*($B$5/(0.05+$B$5))), INT(C9*$B$5))", "메가와리: 없음 / 메가포: 실결제금액 기준 (인당 1만엔 한도 방어 수식 내장)"),
        ("④ 프로모션 시스템 이용료", "=ROUND(B9*0.01, 1)", 0, "메가와리: 고객 최종 실결제금액의 1% 부과 / 메가포: 면제"),
        ("🔥 판매자 총 공제 비용 (JPY)", "=SUM(B10:B13)", "=SUM(C10:C13)", "항목 ① + ② + ③ + ④ 총합"),
        ("최종 정산 공제 비율", "=B14/B8", "=C14/C8", "할인 적용 전 판매가격 대비 총 공제 금액 비율"),
        ("💰 판매자 최종 정산금 (JPY)", "=B9-B14", "=C9-C14", "고객 실결제금액 - 판매자 총 공제 비용"),
        ("▶ 총 공제 비용 (KRW 환산)", "=B14*$B$3", "=C14*$B$3", "엔화 총 공제 비용 × 상단 입력 환율"),
        ("▶ 최종 정산 입금액 (KRW 환산)", "=B16*$B$3", "=C16*$B$3", "엔화 최종 정산금 × 상단 입력 환율")
    ]

    # 7. 엑셀에 데이터 및 수식 주입
    for i, row_content in enumerate(rows_data, 8):
        ws.row_dimensions[i].height = 22
        for j, val in enumerate(row_content, 1):
            cell = ws.cell(row=i, column=j)
            cell.value = val
            cell.font = font_data
            cell.border = thin_border
            
            # 정렬 및 숫자 서식 서싱
            if j == 1:
                cell.alignment = align_left
            elif j in [2, 3]:
                cell.alignment = align_right
                # 퍼센트 서식
                if i == 15: 
                    cell.number_format = '0.0%'
                # 원화 서식
                elif i in [17, 18]:
                    cell.number_format = '"₩"#,##0'
                # 엔화 서식
                else:
                    cell.number_format = '"¥"#,##0'
            else:
                cell.alignment = align_left
                
        # 입력칸 강조 스타일링 (판매가격 입력 부분)
        if i == 8:
            ws.cell(row=i, column=2).fill = fill_input
            ws.cell(row=i, column=3).fill = fill_input
            ws.cell(row=i, column=2).font = font_bold
            ws.cell(row=i, column=3).font = font_bold
            
        # 결과 및 합계행 강조 스타일링
        if i in [14, 16, 18]:
            ws.cell(row=i, column=1).font = font_bold
            ws.cell(row=i, column=2).font = font_bold
            ws.cell(row=i, column=3).font = font_bold
            if i == 18:  # 최종 한국돈 입금액 라인 완전히 강조
                ws.cell(row=i, column=1).fill = fill_total
                ws.cell(row=i, column=2).fill = fill_total
                ws.cell(row=i, column=3).fill = fill_total
                ws.cell(row=i, column=2).font = font_total
                ws.cell(row=i, column=3).font = font_total
                ws.cell(row=i, column=1).border = double_bottom
                ws.cell(row=i, column=2).border = double_bottom
                ws.cell(row=i, column=3).border = double_bottom

    # 8. 열 너비 자동 조절
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 50

    # 9. 파일 저장
    file_name = "Qoo10_Mega_Calculator.xlsx"
    wb.save(file_name)
    print(f"성공적으로 엑셀 서식 파일이 생성되었습니다: {file_name}")

if __name__ == "__main__":
    create_qoo10_excel()
